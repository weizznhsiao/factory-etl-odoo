"""
Core ETL processing logic adapted for in-memory Odoo integration.
Parsers are auto-detected by trying each registered parser in order.
"""

from __future__ import annotations
import io
from typing import Optional
import pandas as pd
import logging

from .mapper import build_horizontal_row, build_size_cols_map
from .registry import try_all_parsers
from . import parsers  # noqa: F401 (triggers registration)

_logger = logging.getLogger(__name__)


def process_orders_in_memory(
    order_files: list[tuple[str, bytes]],
    template_bytes: bytes,
) -> tuple[Optional[bytes], str]:
    """
    Processes a batch of order files using a template.
    Parser is auto-detected by trying all registered parsers until one succeeds.
    Returns a tuple: (generated_excel_bytes, log_summary)
    """
    _logger.info("Starting in-memory ETL process...")

    all_records = []
    processed_count = 0
    skipped_files = []

    for filename, file_bytes in order_files:
        import os
        _, ext = os.path.splitext(filename)
        ext = ext.lower()

        # Only process supported file types
        if ext not in ('.xls', '.xlsx', '.xlsb', '.pdf'):
            _logger.warning(f"Skipping unsupported file type: {filename}")
            skipped_files.append(filename)
            continue

        file_records = try_all_parsers(filename=filename, file_bytes=file_bytes)

        if file_records is None:
            _logger.warning(f"No parser could handle file: {filename}")
            skipped_files.append(filename)
            continue

        all_records.extend(file_records)
        processed_count += 1

    if not all_records:
        skipped_info = f"（跳過：{', '.join(skipped_files)}）" if skipped_files else ""
        return None, f"無法從提供的檔案中解析出任何訂單記錄。{skipped_info}"

    # Load template
    try:
        f_temp = io.BytesIO(template_bytes)
        df_template = pd.read_excel(f_temp, sheet_name=0)
        headers = list(df_template.columns)
    except Exception as e:
        return None, f"Failed to read horizontal template: {str(e)}"

    size_cols_map = build_size_cols_map(headers)

    # Generate rows
    output_rows = [build_horizontal_row(r, headers, size_cols_map) for r in all_records]
    df_output = pd.DataFrame(output_rows, columns=headers)

    # Write output to BytesIO using openpyxl for formatting preservation if possible,
    # but since we are modifying an existing template, we use openpyxl load_workbook
    output_buffer = io.BytesIO()
    try:
        import openpyxl
        f_temp.seek(0)
        wb = openpyxl.load_workbook(f_temp)
        ws = wb.active

        # Clear existing data rows (keep header row 1)
        while ws.max_row > 1:
            ws.delete_rows(2)

        for r_dict in output_rows:
            row_values = [r_dict[col] for col in headers]
            ws.append(row_values)

        wb.save(output_buffer)
        wb.close()
    except Exception as e:
        _logger.error(f"Failed to save using openpyxl, falling back to pandas: {e}")
        # Fallback to pandas
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            df_output.to_excel(writer, index=False)

    skipped_info = f"，跳過 {len(skipped_files)} 個無法辨識的檔案" if skipped_files else ""
    return output_buffer.getvalue(), f"成功處理 {processed_count} 個檔案，共產生 {len(output_rows)} 筆記錄{skipped_info}。"
